#!/usr/bin/env python3
"""
Import a portfolio Excel file into the portfolio tracker API.

Expected workbook format
────────────────────────
Sheet "App"  (required)
  Row 1 — headers:  Period | Income | Expenses | <Asset Name> | <Asset Name> …
  Row 2 — metadata: (blank) | (blank) | (blank) | <currency,type[,excel_currency]> | …
  Row 3+ — data:    2024-01-01 | 9500 | 4200 | 32000 | 15000 | …

  Metadata cell format (comma-separated, case-insensitive):
    <currency>,<type>                 e.g.  AUD,cash   or   USD,equities
    <currency>,<type>,<excel_ccy>     e.g.  USD,equities,AUD
      • currency      — AUD | USD | EUR | GBP  (the asset's native currency)
      • type          — cash | equities | equity | crypto | property | bonds | other
      • excel_currency — currency the Excel column is actually in; omit if same as currency
                         (use this when a USD asset was tracked in AUD in the spreadsheet)

  Leave a cell blank if the asset had no snapshot that month.

Sheet "Transactions"  (optional)
  Row 1 — headers: Date | Description | Amount | Category
  Row 2+ — data

Usage
─────
  python import_excel.py --file Accounting.xlsx [--api http://localhost:8000]
"""

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

import requests
import openpyxl

API = "http://localhost:8000/api"

TYPE_COLORS = {
    "cash":     "#4e8ef7",
    "equities": "#2ec27e",
    "crypto":   "#f5a623",
    "property": "#e05c5c",
    "bonds":    "#b07fd4",
    "other":    "#8b91a8",
}

TYPE_ALIASES = {
    "equity": "equities", "equities": "equities", "stock": "equities", "stocks": "equities",
    "cash": "cash", "savings": "cash", "bank": "cash",
    "crypto": "crypto", "cryptocurrency": "crypto",
    "property": "property", "real estate": "property", "realestate": "property",
    "bonds": "bonds", "bond": "bonds", "fixed income": "bonds",
    "other": "other",
}

FIXED_COLS = {"Period", "Income", "Expenses"}


# ── Helpers ────────────────────────────────────────────────────────────────────

def post(path, body):
    r = requests.post(API + path, json=body, timeout=10)
    if not r.ok:
        print(f"  WARN {r.status_code} {path}: {r.text[:120]}")
    return r.json() if r.ok else None

def get(path):
    r = requests.get(API + path, timeout=10)
    r.raise_for_status()
    return r.json()

def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                pass
    return None

def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def normalize_type(raw: str) -> str:
    return TYPE_ALIASES.get(raw.strip().lower(), "other")

def parse_meta(cell) -> tuple[str, str, str] | None:
    """Parse 'USD,equities' or 'USD,equities,AUD' → (currency, type, excel_currency)."""
    if not cell:
        return None
    parts = [p.strip() for p in str(cell).split(",")]
    if len(parts) < 2:
        return None
    currency = parts[0].upper()
    asset_type = normalize_type(parts[1])
    excel_currency = parts[2].upper() if len(parts) > 2 else currency
    return currency, asset_type, excel_currency


# ── Assets ─────────────────────────────────────────────────────────────────────

def ensure_assets(asset_cols: list[tuple[str, tuple[str, str, str]]]) -> dict[str, int]:
    """Create assets that don't exist yet; return name → id map."""
    existing = {a["name"]: a["id"] for a in get("/assets/?include_inactive=true")}
    ids: dict[str, int] = {}
    for order, (name, (currency, asset_type, _)) in enumerate(asset_cols):
        if name in existing:
            ids[name] = existing[name]
            continue
        color = TYPE_COLORS.get(asset_type, TYPE_COLORS["other"])
        resp = post("/assets/", {
            "name": name, "type": asset_type, "currency": currency,
            "color": color, "display_order": order, "is_active": True,
        })
        if resp:
            ids[name] = resp["id"]
            print(f"  Created: {name} ({currency}, {asset_type}, id={resp['id']})")
        else:
            print(f"  Failed to create: {name}")
    return ids


# ── FX rate cache ──────────────────────────────────────────────────────────────

_fx_cache: dict[str, dict[str, float]] = {}

def fetch_fx(period: date) -> dict[str, float]:
    key = period.isoformat()
    if key not in _fx_cache:
        try:
            _fx_cache[key] = get(f"/fx-rates/?period={key}")
        except Exception:
            _fx_cache[key] = {}
    return _fx_cache[key]


# ── Snapshots sheet ────────────────────────────────────────────────────────────

def import_snapshots(ws):
    print("\nImporting App sheet…")
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        print("  Need at least 3 rows (header, metadata, data). Skipping.")
        return

    headers  = [str(h).strip() if h is not None else "" for h in rows[0]]
    meta_row = rows[1]
    data_rows = rows[2:]

    period_col  = headers.index("Period")
    income_col  = headers.index("Income")   if "Income"   in headers else None
    expense_col = headers.index("Expenses") if "Expenses" in headers else None

    # Build list of (name, (currency, type, excel_currency)) for asset columns only
    asset_cols: list[tuple[str, tuple[str, str, str]]] = []
    for i, h in enumerate(headers):
        if h in FIXED_COLS or not h:
            continue
        meta = parse_meta(meta_row[i])
        if meta is None:
            print(f"  WARN: no metadata for column '{h}' (row 2 cell {i+1}) — skipping column")
            continue
        asset_cols.append((h, meta))

    if not asset_cols:
        print("  No valid asset columns found.")
        return

    print(f"  Asset columns: {[n for n, _ in asset_cols]}")
    asset_ids = ensure_assets(asset_cols)
    imported = 0
    for row in data_rows:
        period_raw = row[period_col]
        d = to_date(period_raw)
        if not d:
            continue
        period = d.replace(day=1)

        income  = to_float(row[income_col])  if income_col  is not None else None
        expense = to_float(row[expense_col]) if expense_col is not None else None
        if income is not None or expense is not None:
            post("/cashflows/", {
                "period":   period.isoformat(),
                "income":   income or 0.0,
                "expenses": abs(expense) if expense else 0.0,
            })

        for name, (currency, _, excel_currency) in asset_cols:
            col_idx = headers.index(name)
            val = to_float(row[col_idx])
            if val is None:
                continue
            aid = asset_ids.get(name)
            if not aid:
                continue

            fx_rate = None
            if currency != "AUD" or excel_currency != "AUD":
                fx_rates = fetch_fx(period)
                fx_rate = fx_rates.get(currency)

            if excel_currency == "AUD":
                # Column is already in AUD; back-calculate native value
                value_aud    = val
                value_native = round(val / fx_rate, 2) if fx_rate else val
            else:
                # Column is in the asset's native currency
                value_native = val
                value_aud    = round(val * fx_rate, 2) if fx_rate else val

            post("/snapshots/", {
                "asset_id":     aid,
                "period":       period.isoformat(),
                "value_native": value_native,
                "value_aud":    value_aud,
                "fx_rate":      round(fx_rate, 6) if fx_rate else None,
            })

        imported += 1

    print(f"  Processed {imported} rows.")


# ── Transactions sheet ─────────────────────────────────────────────────────────

def import_transactions(ws):
    print("\nImporting Transactions sheet…")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    date_col  = headers.index("Date")
    desc_col  = headers.index("Description")
    amt_col   = headers.index("Amount")
    cat_col   = headers.index("Category") if "Category" in headers else None

    bulk = []
    for row in rows[1:]:
        d = to_date(row[date_col])
        amt = to_float(row[amt_col])
        desc = str(row[desc_col]) if row[desc_col] else None
        if not d or amt is None or not desc:
            continue
        cat = str(row[cat_col]) if cat_col is not None and row[cat_col] else None
        bulk.append({"date": d.isoformat(), "description": desc, "amount": amt, "category": cat})
        if len(bulk) >= 200:
            r = requests.post(API + "/transactions/bulk", json=bulk, timeout=30)
            print(f"  Batch {r.status_code} ({len(bulk)} rows)")
            bulk = []

    if bulk:
        r = requests.post(API + "/transactions/bulk", json=bulk, timeout=30)
        print(f"  Final batch {r.status_code} ({len(bulk)} rows)")
    print("  Done.")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default="sample.xlsx")
    parser.add_argument("--api",  default="http://localhost:8000")
    parser.add_argument("--no-transactions", action="store_true")
    args = parser.parse_args()

    global API
    API = args.api.rstrip("/") + "/api"

    xlsx = Path(args.file)
    if not xlsx.exists():
        print(f"File not found: {xlsx}"); sys.exit(1)

    print(f"Loading {xlsx}…")
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    import_snapshots(wb.worksheets[0])

    if not args.no_transactions and "Transactions" in wb.sheetnames:
        import_transactions(wb["Transactions"])

    print("\nReconciling derived transactions…")
    r = requests.post(API + "/snapshots/reconcile-all", timeout=30)
    print(f"  {r.json()}")

    print("\nImport complete.")

if __name__ == "__main__":
    main()
