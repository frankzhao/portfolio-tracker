#!/usr/bin/env python3
"""Generate sample.xlsx with 5 years of synthetic portfolio data."""

import random
import math
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)

# ── Date range ─────────────────────────────────────────────────────────────────

def months_range(start: date, n: int):
    periods = []
    y, m = start.year, start.month
    for _ in range(n):
        periods.append(date(y, m, 1))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return periods

START   = date(2021, 5, 1)
PERIODS = months_range(START, 60)   # May 2021 → Apr 2026

# ── Asset value generators ─────────────────────────────────────────────────────

def bank_walk(start, annual_drift=0.03, vol=0.04, floor=2000):
    v = start
    out = []
    for _ in PERIODS:
        shock = random.gauss(0, vol)
        v = max(floor, v * (1 + annual_drift / 12 + shock / math.sqrt(12)))
        out.append(round(v, 2))
    return out

def equity_walk(start, annual_drift=0.14, vol=0.18, floor=100):
    v = start
    out = []
    for i, _ in enumerate(PERIODS):
        bear = -0.03 if 9 <= i <= 20 else 0.0
        shock = random.gauss(0, vol)
        v = max(floor, v * (1 + (annual_drift + bear * 12) / 12 + shock / math.sqrt(12)))
        out.append(round(v, 2))
    return out

def crypto_walk(start, floor=50):
    params = []
    for i in range(len(PERIODS)):
        if i < 7:
            params.append((0.12, 0.20))
        elif i < 20:
            params.append((-0.10, 0.18))
        elif i < 32:
            params.append((0.06, 0.14))
        elif i < 44:
            params.append((0.10, 0.18))
        else:
            params.append((0.02, 0.22))
    v = start
    out = []
    for mu, sigma in params:
        v = max(floor, v * (1 + mu + random.gauss(0, sigma)))
        out.append(round(v, 2))
    return out

# ── Asset definitions ──────────────────────────────────────────────────────────
# (generator_fn, metadata_string)
# metadata = "currency,type" written into row 2 of the sheet

ASSET_DEFS = {
    #  name        values (native currency)           metadata row 2
    "CBA":       (bank_walk(28_000, 0.04, 0.05),      "AUD,cash"),
    "Westpac":   (bank_walk(14_000, 0.03, 0.06),      "AUD,cash"),
    "ANZ":       (bank_walk( 6_000, 0.02, 0.07),      "AUD,cash"),
    "Stake":     (equity_walk(4_500, 0.16, 0.17),     "USD,equities"),  # values in USD
    "IBKR":      (equity_walk(8_000, 0.14, 0.14),     "USD,equities"),  # values in USD
    "BitX":      (crypto_walk(3_500),                  "AUD,crypto"),
    "CoinVault": (crypto_walk(2_000),                  "AUD,crypto"),
}

# ── Cash flows ─────────────────────────────────────────────────────────────────

def income_series():
    base = 9_200
    return [round(base * (1 + 0.03 * i / 12 + random.gauss(0, 0.04)), 2) for i in range(60)]

def expense_series():
    base = 4_800
    return [round(base * (1 + 0.025 * i / 12 + random.gauss(0, 0.05)), 2) for i in range(60)]

incomes  = income_series()
expenses = expense_series()

# ── Transactions ───────────────────────────────────────────────────────────────

CATEGORIES = {
    "Groceries":     (-180, -60,  8),
    "Dining":        (-120, -20,  6),
    "Transport":     ( -90, -15,  5),
    "Utilities":     (-220, -80,  2),
    "Entertainment": ( -80, -10,  4),
    "Healthcare":    (-200, -30,  2),
    "Shopping":      (-300, -20,  4),
    "Subscriptions": ( -50, -10,  3),
    "Salary":        (5000, 7500,  1),
    "Freelance":     ( 500, 2500,  1),
}

def make_transactions():
    rows = []
    for period in PERIODS:
        for cat, (lo, hi, freq) in CATEGORIES.items():
            for _ in range(freq):
                day = random.randint(1, 28)
                dt  = date(period.year, period.month, day)
                amt = round(random.uniform(lo, hi), 2)
                rows.append((dt, f"{cat} payment", amt, cat))
    rows.sort(key=lambda r: r[0])
    return rows

transactions = make_transactions()

# ── Write workbook ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
hdr_fill  = PatternFill("solid", fgColor="1e2235")
meta_fill = PatternFill("solid", fgColor="2a2f45")

# ── App sheet (snapshots + metadata) ──────────────────────────────────────────
ws = wb.active
ws.title = "App"

asset_names = list(ASSET_DEFS.keys())
header      = ["Period", "Income", "Expenses"] + asset_names

# Row 1 — headers
ws.append(header)
for col, h in enumerate(header, 1):
    cell = ws.cell(1, col)
    cell.font      = Font(bold=True, color="FFFFFF")
    cell.fill      = hdr_fill
    cell.alignment = Alignment(horizontal="center")

# Row 2 — metadata (blank for fixed columns, "currency,type" for assets)
meta_row = ["", "", ""] + [ASSET_DEFS[n][1] for n in asset_names]
ws.append(meta_row)
for col, val in enumerate(meta_row, 1):
    cell = ws.cell(2, col)
    cell.fill      = meta_fill
    cell.font      = Font(italic=True, color="8b91a8")
    cell.alignment = Alignment(horizontal="center")

# Row 3+ — data
for i, period in enumerate(PERIODS):
    row = [period, incomes[i], expenses[i]] + [ASSET_DEFS[n][0][i] for n in asset_names]
    ws.append(row)

ws.column_dimensions["A"].width = 13
for col in range(2, len(header) + 1):
    ws.column_dimensions[get_column_letter(col)].width = 14

# ── Transactions sheet ─────────────────────────────────────────────────────────
wt = wb.create_sheet("Transactions")
wt.append(["Date", "Description", "Amount", "Category"])
for col in range(1, 5):
    cell = wt.cell(1, col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hdr_fill

for dt, desc, amt, cat in transactions:
    wt.append([dt, desc, amt, cat])

wt.column_dimensions["A"].width = 13
wt.column_dimensions["B"].width = 24
wt.column_dimensions["C"].width = 12
wt.column_dimensions["D"].width = 16

# ── Save ───────────────────────────────────────────────────────────────────────
out = Path("sample.xlsx")
wb.save(out)
print(f"Wrote {out} — {len(PERIODS)} months, {len(transactions)} transactions.")
print(f"Assets: {', '.join(f'{n} ({ASSET_DEFS[n][1]})' for n in asset_names)}")
