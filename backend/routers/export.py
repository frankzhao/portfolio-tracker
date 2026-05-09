from io import BytesIO
from datetime import date

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from models import Asset, Snapshot, CashFlow, Transaction

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/xlsx")
def export_xlsx(db: Session = Depends(get_db)):
    assets = (
        db.query(Asset)
        .order_by(Asset.display_order, Asset.id)
        .all()
    )

    snapshots = db.query(Snapshot).order_by(Snapshot.period, Snapshot.asset_id).all()
    cashflows = db.query(CashFlow).order_by(CashFlow.period).all()
    transactions = (
        db.query(Transaction)
        .filter(Transaction.is_derived == False)
        .order_by(Transaction.date)
        .all()
    )

    # Build period → asset_id → value_native index
    snap_index: dict[date, dict[int, float]] = {}
    for s in snapshots:
        snap_index.setdefault(s.period, {})[s.asset_id] = float(s.value_native)

    cf_index: dict[date, CashFlow] = {cf.period: cf for cf in cashflows}

    all_periods = sorted(snap_index.keys() | cf_index.keys())

    wb = openpyxl.Workbook()

    # ── App sheet ─────────────────────────────────────────────────────────────

    ws = wb.active
    assert ws is not None
    ws.title = "App"

    header_font = Font(bold=True)
    meta_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    # Row 1: headers
    headers = ["Period", "Income", "Expenses"] + [a.name for a in assets]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = header_font

    # Row 2: metadata (blank for fixed cols, "currency,type" for asset cols)
    for col in range(1, 4):
        ws.cell(row=2, column=col).fill = meta_fill
    for col, asset in enumerate(assets, 4):
        cell = ws.cell(row=2, column=col, value=f"{asset.currency.value},{asset.type.value}")
        cell.fill = meta_fill
        cell.alignment = Alignment(horizontal="center")

    # Row 3+: data rows
    for row_idx, period in enumerate(all_periods, 3):
        ws.cell(row=row_idx, column=1, value=period)
        cf = cf_index.get(period)
        if cf:
            ws.cell(row=row_idx, column=2, value=float(cf.income))
            ws.cell(row=row_idx, column=3, value=float(cf.expenses))
        snaps_for_period = snap_index.get(period, {})
        for col, asset in enumerate(assets, 4):
            val = snaps_for_period.get(asset.id)
            if val is not None:
                ws.cell(row=row_idx, column=col, value=val)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    for col_idx, asset in enumerate(assets, 4):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(14, len(asset.name) + 2)

    # ── Transactions sheet ────────────────────────────────────────────────────

    wt = wb.create_sheet("Transactions")

    tx_headers = ["Date", "Description", "Amount", "Category"]
    for col, val in enumerate(tx_headers, 1):
        cell = wt.cell(row=1, column=col, value=val)
        cell.font = header_font

    for row_idx, tx in enumerate(transactions, 2):
        wt.cell(row=row_idx, column=1, value=tx.date)
        wt.cell(row=row_idx, column=2, value=tx.description)
        wt.cell(row=row_idx, column=3, value=float(tx.amount))
        wt.cell(row=row_idx, column=4, value=tx.category)

    wt.column_dimensions["A"].width = 12
    wt.column_dimensions["B"].width = 40
    wt.column_dimensions["C"].width = 12
    wt.column_dimensions["D"].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=portfolio_export.xlsx"},
    )
